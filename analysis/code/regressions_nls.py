import numpy as np
import pandas as pd
from scipy.optimize import minimize, least_squares
from scipy.stats import norm
from openpyxl import Workbook, load_workbook, utils
from openpyxl.styles import Alignment, Font
import string


### Functions

def trim(X, percent):
    ''' Trims X with the percent input by finding the top and bottom 
    (percent/2) observations and then returns a matrix of binary variables 
    indicating rows of X that would not be trimmed
    '''
    
    alpha = (1 - percent)/2
    n, k = np.shape(X)
    t_ind = np.zeros((n, k))
    
    for i in range(0, k):
        upper_bd = np.percentile(X[:,i], (1 - alpha)*100)
        lower_bd = np.percentile(X[:,i], (alpha)*100)
        t_ind[:, i] = [int(lower_bd < x < upper_bd) for x in X[:,i]]
        
    return np.prod(t_ind, axis = 1)


def CE_1(Y, X, arg, r):
    ''' Conditional Expectation of Y given X at arg (matrix) with bandwith r 
    using a gaussian kernel 
    '''

    n_arg = np.shape(arg)[0]
    n = np.shape(X)[0]
    h = (n**(-r)) * np.std(X, axis = 0, ddof = 1)
    e = np.zeros((n_arg, 1))
    
    for j in range(0, n_arg):
        k = np.divide(norm.pdf(np.divide((arg[j] - X), h)), h)
        k = np.prod(k, axis = 1)   
        e[j] = (Y.T*k/n)/np.mean(k)
        
    return e


def SLS_1(b, Y, X, X_ind):
    ''' Semiparametric least-squares using CE_1 and bandwith = 1/5
    '''
    
    v = X * np.matrix(b).T
    EY = CE_1(Y, v, v, 1/5)
    residual = np.power((Y - EY), 2)
    
    return (-0.5 * np.matrix(X_ind)*residual)


def run_semiparametric_regression(Y, X, guess, trim_percent = 0.98, 
    xtol = 0.001, maxiter = 1):
    ''' Runs SLS with some default parameters, approximates an initial guess 
    using a least squares optimization routine and then the final estimate 
    using BFGS
    '''

    obj_f = lambda x_0: -SLS_1(np.append(np.array([1]), x_0), 
        Y, X, trim(X, 0.98))[0,0]

    print('    Running LS...')
    result = least_squares(obj_f, list(np.array(guess).flatten()), 
        xtol = xtol)

    print('    BFGS...')
    result = minimize(obj_f, result.x, method='BFGS', 
        options = {'maxiter': maxiter})

    return result


def convert_hessian_to_cov(Y, X, results):
    ''' Converts the result output from scipy.optimize into a covariance matrix 
    by taking the inverse of the hessian and multiplying by an estimate of the 
    variance of the residuals 
    '''
    
    sigma_2_hat = np.mean(np.power(Y - X*np.matrix(results.x).T, 2))
    return results.hess_inv * sigma_2_hat


def compute_marginal_effect(Y, X, ind, point, beta, delta = 0.01):
    ''' Finds the marginal effects at a given point using CE_1, delta 
    represents the amount to nudge the point by when calculating the marginal 
    effects, ind refers to the index of the variable being nudged
    '''

    point_nudge = np.copy(point)
    point_nudge[0, ind] = point_nudge[0, ind] + delta
    point_nudge = np.matrix(point_nudge)
    
    v_hat = X*beta
    v_hat_avg = point*beta
    v_hat_avg_nudge = point_nudge*beta
    
    return np.asscalar(CE_1(Y, v_hat, v_hat_avg_nudge, 1/5) - 
        CE_1(Y, v_hat, v_hat_avg, 1/5))/delta


def find_tstats(Y, X, results):
    ''' Computes t_stats using an input of variables and the results 
    of a scipy.minimize routine (must output a hessian)
    '''
    
    V = convert_hessian_to_cov(Y, X, results)

    n = np.shape(results.x)[0]
    theta = results.x/results.x[0]
    t_stats = np.zeros(shape = (n))
    t_stats[0] = np.nan # first t-stat is unknown

    for i in range(1, n):
        t_stats[i] = theta[i] / np.sqrt(V[i,i])

    return t_stats    

def get_sig_stars(coeff, stderr, p_value_labels):
    ''' Outputs significance stars after calculating the t-stat 
    '''
    t_stat = coeff/stderr
    p_val  = 2*(1 - norm.cdf(np.abs(t_stat)))
    below_ind = np.where([p_val < x for x in p_value_labels.keys()])[0]
    below_vals = [list(p_value_labels.keys())[i] for i in below_ind]
    if not below_vals:
        return ''
    else:
        min_p_val = np.min(below_vals)
        return p_value_labels[min_p_val]



### Params

data_dmd_loc = '../../data/processed/regression_data_levels_demeaned.csv'

output_file_loc  = '../results/semiparametric_regressions_'
output_file_name = input('Enter file name: ')
workbook_file_loc = output_file_loc + output_file_name

sample_query = input('Enter data query: ')

# 'nudge length' for calculating marginal effects
delta = .001

# params for optimization routines
reg_xtol = 0.01
reg_maxiter = 10

p_value_labels = {0.05: '*', 0.01: '**', 0.001: '***'}



### Main

## Load data

data_df = pd.read_csv(data_dmd_loc).dropna()
data_df['PrImp_Pct_Rebate_Dummy'] = data_df['PrImp_Pct'] * \
                                        data_df['Rebate_Dummy']
data_df['PrImp_AvgAmt_Rebate_Dummy'] = data_df['PrImp_AvgAmt'] * \
                                         data_df['Rebate_Dummy']
data_df['PrImp_ExpAmt_Rebate_Dummy'] = data_df['PrImp_ExpAmt'] * \
                                         data_df['Rebate_Dummy']
data_df['PrImp_AvgT_Rebate_Dummy'] = data_df['PrImp_AvgT'] * \
                                         data_df['Rebate_Dummy']
data_df['All_AvgT_Rebate_Dummy'] = data_df['All_AvgT'] * \
                                         data_df['Rebate_Dummy']

## Information about sample

sample_frac = 1 # None => All obs
data_df = data_df.sample(frac = sample_frac).query(sample_query)

print('Market Centers: ', end = '') 
print(data_df['MarketCenter'].unique())

print('Brokers: ', end = '') 
print(data_df['Broker'].unique())

print('Exchanges: ', end = '') 
print(data_df['Exchange'].unique())

print('OrderTypes: ', end = '') 
print(data_df['OrderType'].unique())

print('Samples: %d' % data_df.shape[0])
print('Sparsity: %0.2f%%' % 
    (100*data_df.query('MktShare == 0').shape[0] / data_df.shape[0]))


## Regressions

# Fits
fit1_formula = 'MktShare ~ PrImp_Pct + PrImp_AvgAmt + PrImp_AvgT'
fit2_formula = 'MktShare ~ PrImp_ExpAmt + PrImp_AvgT'
fit3_formula = 'MktShare ~ PrImp_Pct + PrImp_AvgAmt + All_AvgT'
fit4_formula = 'MktShare ~ PrImp_ExpAmt + All_AvgT'

formulaCols = lambda x: x.replace(' ', '').replace('~', '+').split('+') 
fit_formulae = [fit1_formula, fit2_formula, fit3_formula, fit4_formula]
fit_formulae = [formulaCols(x) for x in fit_formulae]

# Store results
fit_results = [None] * len(fit_formulae)


## Get Results
for i in range(0, len(fit_formulae)):
    
    ## Get results
    print('Regressing with fit %d...' % i )
    
    data = data_df[fit_formulae[i]]
    X = np.matrix(data)[:, 1:]
    Y = np.matrix(data)[:, 0]

    guess = X[1, 1:]
    results = run_semiparametric_regression(Y, X, guess, 
        xtol = reg_xtol, maxiter = reg_maxiter)
    
    fit_results[i] = results
    
    ## Update results
    # Normalize results with first coefficient
    fit_results[i].x = np.append(1, fit_results[i].x)
    
    # Add dictionary of coefficients 
    fit_results[i].coeffs  = {fit_formulae[i][1:][j]: fit_results[i].x[j] for 
        j in range(0, len(fit_formulae[i])-1)}
    
    # Add dictionary of standard errors 
    V = convert_hessian_to_cov(Y, X, fit_results[i])
    fit_results[i].stderrs = {fit_formulae[i][1:][j]: np.sqrt(V[j-1,j-1]) for 
        j in range(1, len(fit_formulae[i])-1)}
    fit_results[i].stderrs[fit_formulae[i][1:][0]] = np.nan
    
print('Complete')  


## Marginal Effects

for i in range(0, len(fit_formulae)):
    
    data = data_df[fit_formulae[i]]
    X = np.matrix(data)[:, 1:]
    Y = np.matrix(data)[:, 0]
    
    fit_results[i].marginal_effects = {}
    
    for j in range(0, len(fit_results[i].x)):  
        
        temp_dict = {}
        
        for percentile in range(20, 81, 20):

            X_percentile = np.percentile(X, percentile, axis = 0)
            
            temp_dict[percentile] = compute_marginal_effect(Y, X, j, 
                np.matrix(X_percentile), np.matrix(fit_results[i].x).T, 
                delta = delta)
            
        fit_results[i].marginal_effects[fit_formulae[i][j+1]] = temp_dict


## Export Results

# Open workbook
wb = Workbook()


# Summary Page
ws = wb.create_sheet(title = 'Summary')
ws['B2'] = 'n'
ws['B3'] = data_df.shape[0]


# Coefficient Results Page
ws = wb.create_sheet(title = 'Coefficient Results')
ws.column_dimensions["B"].width = 30


## Label regressors 

# Get regressors besides for market share
fit_regressors = sorted(list(set(sum(fit_formulae, []))-set(['MktShare'])))
fit_regressors

regressor_cells = {}

# First label's row
regressor_label_row = 4

for i in range(0, len(fit_regressors)):

    cell = 'B' + str(regressor_label_row)
   
    ws[cell] = fit_regressors[i]
    ws[cell].alignment = Alignment(horizontal = 'right')
    ws[cell].font = Font(bold = True)
    
    regressor_cells[fit_regressors[i]] = regressor_label_row
    regressor_label_row += 2


## Label regressand

start_cell = 'C2'
end_cell   = string.ascii_uppercase[2*len(fit_results)] + '2'

ws.merge_cells(start_cell + ':' + end_cell)

ws[start_cell] = 'MktShare'
ws[start_cell].alignment = Alignment(horizontal = 'center')
ws[start_cell].font = Font(bold = True)


## Label fits

for i in range(0, len(fit_results)):
    
    cell_row = 3
    cell_col = string.ascii_uppercase[2*i + 2]
    cell = cell_col + str(cell_row)
    
    ws[cell] = 'Fit ' + str(i+1)
    ws[cell].alignment = Alignment(horizontal = 'center')
    ws[cell].font = Font(underline = 'single')
    
    # adjust cell widths
    ws.column_dimensions[cell_col].width = 15
    ws.column_dimensions[string.ascii_uppercase[2*i + 3]].width = 5


## Enter results

for i in range(0, len(fit_results)):
    
    fit_column = string.ascii_uppercase[2*i + 2]
    
    for regressor in fit_results[i].coeffs.keys():
        
        coeff = fit_results[i].coeffs[regressor]
        stderr = fit_results[i].stderrs[regressor]
        
        regressor_label_row = regressor_cells[regressor]
        cell = fit_column + str(regressor_label_row)
        cell_below = fit_column + str(regressor_label_row + 1)
        
        coeff = str(np.round(coeff, decimals = 4)) + get_sig_stars(coeff, 
            stderr, p_value_labels)
        
        ws[cell] = coeff
        ws[cell].alignment = Alignment(horizontal = 'center')
        
        ws[cell_below] = stderr
        ws[cell_below].alignment = Alignment(horizontal = 'center')


## Marginal Effects

for i in range(0, len(fit_results)):
    
    ws = wb.create_sheet(title = 'Fit ' + str(i+1) +' Marginal Effects')
    ws.column_dimensions["B"].width = 30
    
    fit_i_regressors = list(fit_results[i].marginal_effects.keys())
    percentiles = sorted(list((next(
        iter(fit_results[i].marginal_effects.values()))).keys()))
    
    ## Label regressors
    
    # First label's row
    regressor_label_row = 4
    
    regressor_cells = {}

    for j in range(0, len(fit_i_regressors)):

        cell = 'B' + str(regressor_label_row)

        ws[cell] = fit_i_regressors[j]
        ws[cell].alignment = Alignment(horizontal = 'right')
        ws[cell].font = Font(bold = True)

        regressor_cells[fit_i_regressors[j]] = regressor_label_row
        regressor_label_row += 2
        
    ## Table title label
    start_cell = 'C2'
    end_cell   = string.ascii_uppercase[2*len(percentiles)] + '2'

    ws.merge_cells(start_cell + ':' + end_cell)

    ws[start_cell] = 'Marginal Effects (MktShare)'
    ws[start_cell].alignment = Alignment(horizontal = 'center')
    ws[start_cell].font = Font(bold = True)
    
    ## Label percentiles
    for j in range(0, len(percentiles)):
        
        cell_row = 3
        cell_col = string.ascii_uppercase[2*j + 2]
        cell = cell_col + str(cell_row)

        ws[cell] = 'Pct ' + str(percentiles[j])
        ws[cell].alignment = Alignment(horizontal = 'center')
        ws[cell].font = Font(underline = 'single')
        
        # adjust cell widths
        ws.column_dimensions[cell_col].width = 15
        ws.column_dimensions[string.ascii_uppercase[2*j + 3]].width = 5

    ## Enter results
    for regressor in fit_results[i].marginal_effects.keys():
    
        me_pct_dict = fit_results[i].marginal_effects[regressor]
        
        for k in range(0, len(percentiles)):

            pct_col = string.ascii_uppercase[2*k + 2]    
            cell = pct_col + str(regressor_cells[regressor])

            ws[cell] = np.round(
                fit_results[i].marginal_effects[regressor][percentiles[k]], 
                decimals = 6)
            ws[cell].alignment = Alignment(horizontal = 'center')


## Save and Close
del wb['Sheet']
wb.save(workbook_file_loc)
wb.close()